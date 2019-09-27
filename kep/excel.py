import re
import pandas as pd

from openpyxl.utils.cell import coordinate_to_tuple

# string operations


def uncomment_string(s: str):
    """See https://regex101.com/r/LdAJBH/1"""
    pat = r"([\d,.]*)(?:\d+\))*$"
    return re.search(pattern=pat, string=s).group(1).replace(',', '.')


assert uncomment_string('244873)') == '24487'


def uncomment(x, type_):
    if isinstance(x, str):
        x = uncomment_string(x)
    return type_(x)

# Excel operations


def first_year(wb, v):
    row, _ = coordinate_to_tuple(v.cell)
    return wb[v.sheet].cell(row=row, column=1).value


def years(wb, v) -> [int]:
    row, _ = coordinate_to_tuple(v.cell)
    ws = wb[v.sheet]
    result = []
    while True:
        value = ws.cell(row=row, column=1).value
        try:
            year = uncomment(value, int)
        except (ValueError, TypeError):
            break
        result.append(year)
        row += 1
    return result


def width(freq):
    return dict(A=1, Q=4, M=12)[freq]


def range_values(wb, v):
    row0, col0 = coordinate_to_tuple(v.cell)
    ws = wb[v.sheet]
    for row in ws.iter_rows(min_row=row0,
                            max_row=row0 + len(years(wb, v)) - 1,
                            min_col=col0,
                            max_col=col0 + width(v.freq) - 1):
        for cell in row:
            yield cell.value


def to_series(wb, v):
    data = [uncomment(x, float) for x in range_values(wb, v) if x]
    start_year = first_year(wb, v)
    ix = pd.date_range(start=f'1/1/{start_year}',
                             freq=v.freq,
                             periods=len(data))
    ts = pd.Series(data=data, index=ix)
    ts.name = v.name
    return ts
